import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np
import os
import sys
import json
import shutil

def write_excel(analyze_name_list, model_name_list, saved_result_dir, save_excel_file_path, 
                is_paper=False, is_clustering=False, is_eval_mask=False, is_jae=False):
    eval_results_list = []
    for analyze_name, model_name in zip(analyze_name_list, model_name_list):
        print(f'========{analyze_name}')
        json_file_path_gar = os.path.join(saved_result_dir, analyze_name, 'eval_gar_metrics.json')
        if is_eval_mask:
            eval_mask_num = int(model_name.split(' ')[-1])
            if (eval_mask_num!=0):
                json_file_path_gar = os.path.join(saved_result_dir, analyze_name, f'eval_gar_metrics_{eval_mask_num}.json')
        with open(json_file_path_gar, 'r') as f:
            eval_results_dic_gar = json.load(f)

        json_file_path_iar = os.path.join(saved_result_dir, analyze_name, 'eval_iar_metrics.json')
        if os.path.exists(json_file_path_iar):
            with open(json_file_path_iar, 'r') as f:
                eval_results_dic_iar = json.load(f)
        else:
            eval_results_dic_iar = {'IAR MCA':0, 'IAR PMCA':0}
        
        if is_jae:
            json_file_path_jae = os.path.join(saved_result_dir, analyze_name, 'eval_jae_metrics.json')
            with open(json_file_path_jae, 'r') as f:
                eval_results_dic_jae = json.load(f)
        else:
            eval_results_dic_jae = {'JA distance':0}

        eval_results_dic = {**eval_results_dic_gar, **eval_results_dic_iar, **eval_results_dic_jae}

        if is_paper:
            eval_results_dic_update = update_eval_dic_paper(eval_results_dic, analyze_name, model_name)
        else:
            eval_results_dic_update = update_eval_dic(eval_results_dic)
        
        if is_clustering:
            eval_results_dic_update = update_eval_dic_clustering(eval_results_dic, analyze_name, model_name)
        
        if is_jae:
            eval_results_dic_update = update_eval_dic_jae(eval_results_dic, analyze_name, model_name)

        eval_results_list.append(list(eval_results_dic_update.values()))
        eval_metrics_list = list(eval_results_dic_update.keys())

    eval_results_array = np.array(eval_results_list)
    df_eval_results = pd.DataFrame(eval_results_array, model_name_list, eval_metrics_list)
    df_eval_results.to_excel(save_excel_file_path, sheet_name='all')

def refine_excel(save_excel_file_path, is_paper=False, is_clustering=False, is_eval_mask=False, is_jae=False):
    wb = openpyxl.load_workbook(save_excel_file_path)
    ws = wb['all']
    for col_idx, col in enumerate(ws.iter_cols()):
        if col_idx == 0:
            pass
        else:
            row_max_value = max([cell.value for cell in col if type(cell.value) is not str])
            row_min_value = min([cell.value for cell in col if type(cell.value) is not str])
            for row_idx, cell in enumerate(col):
                if is_clustering:
                    cell.number_format = '0.00'
                elif is_jae:
                    cell.number_format = '0.000'
                else:
                    cell.number_format = '0.0'
                if cell.value == row_max_value:
                    if is_paper and not is_jae:
                        cell.fill = PatternFill(fgColor='FFFF00', bgColor="FFFF00", fill_type = "solid")
                        if is_clustering:
                            cell.value = r'\red' + '{' + f'{cell.value:.2f}' + '}'
                        else:
                            cell.value = r'\red' + '{' + f'{cell.value:.1f}' + '}'
                if cell.value == row_min_value:
                    if is_paper and is_jae:
                        cell.fill = PatternFill(fgColor='FFFF00', bgColor="FFFF00", fill_type = "solid")
                        cell.value = r'\red' + '{' + f'{cell.value:.3f}' + '}'

    wb.save(save_excel_file_path)

def update_eval_dic(eval_results_dic):
    eval_results_dic_update = {}
    for eval_met in ['action iou jaccard', 'action iou dice', 'action iou simpson',  'action iou tfidf', 'group activity']:
        for mode in ['people', 'scene']:
            hid_idx_max = 3
            for hit_idx in range(1, hid_idx_max):
                eval_results_dic_update[f'GAR Hit@{hit_idx} ({eval_met}) ({mode})'] = eval_results_dic[f'GAR accuracy hit@{hit_idx} ({eval_met}) ({mode})'] * 100
        for mode in ['people', 'scene']:
            if eval_met != 'group activity':
                eval_results_dic_update[f'GAR mAP rank ({eval_met}) ({mode})'] = eval_results_dic[f'mAP rank ({eval_met}) ({mode})'] * 100

    eval_results_dic_update['IAR MCA'] = eval_results_dic['IAR MCA'] * 100
    eval_results_dic_update['IAR MPCA'] = eval_results_dic['IAR MPCA'] * 100

    return eval_results_dic_update

def update_eval_dic_paper(eval_results_dic, analyze_name, model_name):
    eval_results_dic_update = {}
    for eval_met in ['action iou jaccard', 'action iou tfidf', 'group activity']:
        for mode in ['people', 'scene']:
            if not 'ours' in analyze_name:
                mode = 'people'
            else:
                if 'ind' in model_name:
                    mode = 'people'
                else:
                    mode = 'scene'
        
            hid_idx_max = 4
            # hid_idx_max = 3
            # hid_idx_max = 2
            for hit_idx in range(1, hid_idx_max):
                eval_results_dic_update[f'Hit@{hit_idx} ({eval_met}) ({mode})'] = eval_results_dic[f'GAR accuracy hit@{hit_idx} ({eval_met}) ({mode})'] * 100
                # eval_results_dic_update[f'hit@{hit_idx} ({eval_met})'] = eval_results_dic[f'GAR accuracy hit@{hit_idx} ({eval_met}) ({mode})'] * 100
            if eval_met != 'group activity':
                eval_results_dic_update[f'mAP rank ({eval_met}) ({mode})'] = eval_results_dic[f'mAP rank ({eval_met}) ({mode})'] * 100
                # eval_results_dic_update[f'mAP rank ({eval_met})'] = eval_results_dic[f'mAP rank ({eval_met}) ({mode})'] * 100

    # eval_results_dic_update['IAR MCA'] = eval_results_dic['IAR MCA'] * 100
    # eval_results_dic_update['IAR MPCA'] = eval_results_dic['IAR MPCA'] * 100

    return eval_results_dic_update

def update_eval_dic_clustering(eval_results_dic, analyze_name, model_name):
    eval_results_dic_update = {}
    for eval_met in ['NRI', 'ARI']:
        for mode in ['people', 'scene']:
            if not 'ours' in analyze_name:
                mode = 'people'
            else:
                if 'ind' in model_name:
                    mode = 'people'
                else:
                    mode = 'scene'

            cluster_list = [8, 16, 32]
            for cluster_idx in cluster_list:
                eval_results_dic_update[f'{eval_met} ({mode}) k={cluster_idx}'] = eval_results_dic[f'{eval_met} ({mode}) k={cluster_idx}']

    return eval_results_dic_update

def update_eval_dic_jae(eval_results_dic, analyze_name, model_name):
    eval_results_dic_update = {}
    ja_key = f'JA distance'
    print(eval_results_dic[ja_key], model_name, analyze_name)
    eval_results_dic_update[ja_key] = eval_results_dic[ja_key] * 600

    return eval_results_dic_update

def collect_visualization_results(save_vis_dir, analyze_name_list, model_name_list, saved_result_dir):
    """_summary_
    Args:
        save_vis_dir (str): Directory for saving visualization results
        analyze_name_list (list): List of analyze names
        model_name_list (list): List of model names for our understanding
        saved_result_dir (str): Direcoty in which results are saved
    """

    if not os.path.exists(save_vis_dir):
        os.makedirs(save_vis_dir)
    for analyze_name, model_name in zip(analyze_name_list, model_name_list):
        saved_result_analysis_dir = os.path.join(saved_result_dir, analyze_name)
        for mode in ['people', 'scene']:
            if 'CAD' in analyze_name:
                mode = 'people'
            else:
                if not 'ours' in analyze_name:
                    mode = 'people'
                else:
                    mode = 'scene'

            cm_file_path = os.path.join(saved_result_analysis_dir, f'confusion_matrix_gar_{mode}.png')
            cm_save_path = os.path.join(save_vis_dir, f'cm_{model_name}_{mode}.png')
            shutil.copy(cm_file_path, cm_save_path)
            tsne_file_wo_legend_path = os.path.join(saved_result_analysis_dir, f'tsne_{mode}_wo_legend.png')
            tsne_save_wo_legend_path = os.path.join(save_vis_dir, f'tsne_{model_name}_{mode}_wo_legend.png')
            shutil.copy(tsne_file_wo_legend_path, tsne_save_wo_legend_path)
            tsne_file_path = os.path.join(saved_result_analysis_dir, f'tsne_{mode}.png')
            tsne_save_path = os.path.join(save_vis_dir, f'tsne_legend.png')
            shutil.copy(tsne_file_path, tsne_save_path)

def collect_nn_results(save_nn_dir, analyze_name_list, model_name_list, saved_result_dir):
    """_summary_
    Args:
        save_nn_dir (str): Directory for saving visualization results
        analyze_name_list (list): List of analyze names
        model_name_list (list): List of model names for our understanding
        saved_result_dir (str): Direcoty in which results are saved
    """

    if not os.path.exists(save_nn_dir):
        os.makedirs(save_nn_dir)
    for analyze_name, model_name in zip(analyze_name_list, model_name_list):
        saved_result_analysis_dir = os.path.join(saved_result_dir, analyze_name)
        for mode in ['people', 'scene']:
            if 'CAD' in analyze_name:
                mode = 'people'
            else:
                if not 'ours' in analyze_name:
                    mode = 'people'
                else:
                    mode = 'scene'

            nn_file_path = os.path.join(saved_result_analysis_dir, f'nn_video_{mode}.xlsx')
            nn_save_path = os.path.join(save_nn_dir, f'nn_video_{model_name}_{mode}.xlsx')
            shutil.copy(nn_file_path, nn_save_path)

def refine_config(cfg):
    if 'use_recon_loss' in dir(cfg):
        pass
    else:
        cfg.use_recon_loss = False

    if 'use_ind_feat' in dir(cfg):
        pass
    else:
        cfg.use_ind_feat = 'loc_and_app'

    if 'use_ind_feat_crop' in dir(cfg):
        pass
    else:
        cfg.use_ind_feat_crop = 'roi_multi'

    if 'person_size' in dir(cfg):
        pass
    else:
        cfg.person_size = 224, 224

    if 'trans_head_num' in dir(cfg):
        pass
    else:
        cfg.trans_head_num = 1

    if 'trans_layer_num' in dir(cfg):
        pass
    else:
        cfg.trans_layer_num = 1

    if 'final_head_mid_num' in dir(cfg):
        pass
    else:
        cfg.final_head_mid_num = 2

    return cfg

def count_parameters(model):
    return sum(p.numel() for p in model.parameters() if p.requires_grad)